"""
================================================================================
highlighting_service.py - 証跡ハイライトサービス
================================================================================

【概要】
監査結果に基づいて証跡ファイルの該当箇所をハイライトします。
以下の3つのモードで動作します：
1. PDFハイライト: 既存のPDFファイル内のテキストをハイライト
2. Excelハイライト: Excelファイル内のセルをハイライト
3. テキスト変換PDF: Word/Email/CSV/画像などのテキストをPDF化してハイライト

【依存ライブラリ】
- fitz (PyMuPDF): PDF操作
- openpyxl: Excel操作
- reportlab: PDF生成

【使用例】
```python
service = HighlightingService()
highlighted_files = await service.highlight_evidence(audit_result, context)
```
"""
import os
import logging
import base64
import tempfile
import re
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

# 外部ライブラリのインポート（利用不可時のハンドリング付き）
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    openpyxl = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import yellow, black
except ImportError:
    canvas = None

# from core.tasks.base_task import AuditContext, AuditResult, EvidenceFile
# 循環参照とパッケージ初期化の問題を回避するため core.types を使用
from core.types import AuditContext, AuditResult, EvidenceFile

logger = logging.getLogger(__name__)


class HighlightingService:
    """
    証跡ハイライト機能を提供するサービスクラス
    """

    def __init__(self, output_dir: str = "highlighted_evidence"):
        """
        初期化

        Args:
            output_dir (str): ハイライト済みファイルの保存先ディレクトリ名
                              （コンテキストのevidence_linkからの相対パス）
        """
        self.output_dir_name = output_dir
        self._check_dependencies()

    def _check_dependencies(self):
        """依存ライブラリの確認"""
        missing = []
        if not fitz:
            missing.append("pymupdf")
        if not openpyxl:
            missing.append("openpyxl")
        if not canvas:
            missing.append("reportlab")

        if missing:
            logger.warning(f"以下のライブラリが不足しています: {', '.join(missing)}。一部機能が制限されます。")
        else:
            logger.debug("ハイライト用ライブラリは全て利用可能です")

    def _normalize_text(self, text: str) -> str:
        """
        テキストの正規化（空白・改行の削除、全角・半角の統一）
        """
        if not text:
            return ""

        import unicodedata
        # NFKC正規化（全角英数→半角、半角カナ→全角など）
        text = unicodedata.normalize('NFKC', text)
        # 改行と空白を削除（厳密な一致ではなく、内容の一致を見るため）
        text = re.sub(r'[\s\u3000]+', '', text)
        return text

    async def highlight_evidence(self, audit_result: AuditResult, context: AuditContext) -> List[Dict[str, str]]:
        """
        監査結果に基づいて証跡ファイルをハイライト
        """
        if not audit_result.document_reference:
            logger.info("引用箇所がないためハイライトをスキップします")
            return []

        logger.info(f"ハイライト処理開始: {len(context.evidence_files)}ファイルの証跡")

        # 引用箇所を解析
        quotes_by_file = self._parse_quotes(audit_result.document_reference)
        logger.debug(f"抽出された引用: {list(quotes_by_file.keys())}")

        highlighted_files_info = []

        # 保存先ディレクトリの決定
        base_dir = Path(context.evidence_link) if context.evidence_link else Path(tempfile.gettempdir())
        save_dir = base_dir / self.output_dir_name

        try:
            save_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.warning(f"保存先ディレクトリ作成失敗 ({save_dir}): {e} -> 一時ディレクトリを使用します")
            save_dir = Path(tempfile.gettempdir()) / "ic_audit_highlighted"
            save_dir.mkdir(parents=True, exist_ok=True)

        for ef in context.evidence_files:
            try:
                # このファイルに関連する引用があるか確認
                file_quotes = []

                # ファイル名マッチングの強化 (正規化して比較)
                ef_stem = Stem(ef.file_name)
                ef_stem_norm = self._normalize_text(ef_stem)

                for fname, quotes in quotes_by_file.items():
                    # 1. 完全一致
                    if fname == ef.file_name:
                        file_quotes.extend(quotes)
                        continue

                    # 2. 部分一致 (相互包含)
                    if fname in ef.file_name or ef.file_name in fname:
                        file_quotes.extend(quotes)
                        continue

                    # 3. 拡張子なしでの一致
                    fname_stem = os.path.splitext(fname)[0]
                    if fname_stem == ef_stem:
                        file_quotes.extend(quotes)
                        continue

                    # 4. 正規化比較
                    fname_norm = self._normalize_text(fname)
                    if fname_norm in ef_stem_norm or ef_stem_norm in fname_norm:
                        file_quotes.extend(quotes)

                # 重複排除
                file_quotes = list(set(file_quotes))

                logger.info(f"ファイル処理: {ef.file_name} (抽出引用数: {len(file_quotes)})")

                output_filename = f"highlighted_{ef.file_name}"
                ext = ef.extension.lower()
                processed = False
                output_path = save_dir / output_filename

                # Base64デコードして一時ファイルに保存
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
                    tmp_file.write(base64.b64decode(ef.base64_content))
                    tmp_path = tmp_file.name

                if ext == '.pdf':
                    if fitz:
                        processed = self._highlight_pdf(tmp_path, file_quotes, str(output_path))
                    else:
                        processed = self._copy_file(tmp_path, output_path)

                elif ext in ['.xlsx', '.xls']:
                    if openpyxl:
                        if ext == '.xlsx':
                            processed = self._highlight_excel(tmp_path, file_quotes, str(output_path))
                        else:
                            logger.info(f"Excel (.xls) は直接ハイライト非対応のため、コピーのみを行います: {ef.file_name}")
                            processed = self._copy_file(tmp_path, output_path)
                    else:
                        processed = self._copy_file(tmp_path, output_path)

                else:
                    # その他の形式: テキストPDF化
                    content_text = ""
                    try:
                        with open(tmp_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content_text = f.read()
                    except Exception:
                        content_text = "(バイナリファイルのためテキスト表示不可)"

                    if canvas:
                        processed = self._generate_highlighted_pdf_from_text(
                            content_text, file_quotes, str(output_path), original_filename=ef.file_name
                        )
                    else:
                        # reportlabがない場合はコピー（拡張子そのまま）
                         processed = self._copy_file(tmp_path, output_path)

                # 一時ファイル削除
                try:
                    os.remove(tmp_path)
                except:
                    pass

                if processed:
                    # ファイルの内容をBase64で読み込む（クライアント側で保存するため）
                    try:
                        with open(output_path, "rb") as f:
                            file_content = f.read()
                            b64_content = base64.b64encode(file_content).decode("utf-8")
                    except Exception as e:
                        logger.error(f"ハイライト済みファイルの読み込み失敗: {e}")
                        b64_content = ""

                    highlighted_files_info.append({
                        "fileName": output_path.name,
                        "filePath": str(output_path.parent),
                        "base64": b64_content  # 追記: Base64エンコードされたコンテンツ
                    })
                    logger.info(f"ハイライト済みファイル生成完了: {output_path.name}")

            except Exception as e:
                logger.error(f"ファイル処理エラー ({ef.file_name}): {e}", exc_info=True)

        return highlighted_files_info

    def _copy_file(self, src: str, dst: Path) -> bool:
        """ファイルをコピーする（ハイライトなしフォールバック）"""
        try:
            import shutil
            shutil.copy(src, dst)
            logger.debug(f"ファイルコピー完了（ハイライトなし）: {dst.name}")
            return True
        except Exception as e:
            logger.error(f"ファイルコピー失敗: {e}")
            return False

    def _parse_quotes(self, document_reference: str) -> Dict[str, List[str]]:
        """
        audit_result.document_reference からファイルごとの引用を抽出
        """
        quotes = {}
        if not document_reference:
            return quotes

        text = document_reference.replace('\r\n', '\n')
        current_file = "unknown"

        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue

            # [filename] または 【filename】 の検出
            match = re.match(r'^[\[【](.*?)[\]】]', line)
            if match:
                current_file = match.group(1).strip()
                content = line[match.end():].strip()
                # プレフィックス除去
                content = re.sub(r'^(page|p\.?)\s*\d+.*?[：:]', '', content, flags=re.IGNORECASE).strip()
                content = re.sub(r'^.*?[：:]', '', content).strip()

                if content:
                    if current_file not in quotes:
                        quotes[current_file] = []
                    quotes[current_file].append(content)
            else:
                if current_file in quotes and current_file != "unknown":
                    quotes[current_file].append(line)

        return quotes

    def _highlight_pdf(self, input_path: str, quotes: List[str], output_path: str) -> bool:
        """PyMuPDFを使用してPDFをハイライト"""
        try:
            doc = fitz.open(input_path)
            found_count = 0

            # 正規化された引用リストを作成
            normalized_quotes = [(q, self._normalize_text(q)) for q in quotes if q and len(q) >= 3]

            for page_num, page in enumerate(doc):
                # ページ全体のテキストを取得（正規化比較用）
                page_text = page.get_text()
                page_text_norm = self._normalize_text(page_text)

                for quote, quote_norm in normalized_quotes:
                    # 1. 通常検索 (PyMuPDFの機能)
                    text_instances = page.search_for(quote)

                    # 2. 正規化マッチング（通常検索で見つからない場合）
                    if not text_instances and quote_norm in page_text_norm:
                        # 簡易的に、ページ全体に注釈をつけるか、ログに出す
                        # 位置特定は難しいので、ここでは左上にメモとして追加する等の代替手段も検討できるが、
                        # 今回は検索できた場合のみハイライトする方針とする。
                        # ただし、ログには出す。
                        logger.debug(f"[PDF] ページ{page_num+1}: 正規化マッチ成功だが位置特定不可: {quote[:20]}...")

                    if text_instances:
                        for inst in text_instances:
                            highlight = page.add_highlight_annot(inst)
                            highlight.set_colors(stroke=(1, 1, 0)) # 黄色
                            highlight.update()
                        found_count += 1

            doc.save(output_path)
            doc.close()

            logger.info(f"[PDF] ハイライト完了: {found_count}箇所をマーク")
            return True

        except Exception as e:
            logger.error(f"PDFハイライトエラー: {e}")
            return False

    def _highlight_excel(self, input_path: str, quotes: List[str], output_path: str) -> bool:
        """openpyxlを使用してExcelセルをハイライト"""
        try:
            wb = openpyxl.load_workbook(input_path)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            found_count = 0

            # 正規化された引用リスト
            normalized_quotes = [self._normalize_text(q) for q in quotes if q and len(q) >= 3]

            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = str(cell.value)
                            cell_text_norm = self._normalize_text(cell_text)

                            matched = False
                            # 1. 通常の部分一致
                            for quote in quotes:
                                if quote in cell_text:
                                    matched = True
                                    break

                            # 2. 正規化部分一致
                            if not matched:
                                for quote_norm in normalized_quotes:
                                    if quote_norm in cell_text_norm:
                                        matched = True
                                        break

                            if matched:
                                cell.fill = yellow_fill
                                found_count += 1

            wb.save(output_path)
            logger.info(f"[Excel] ハイライト完了: {found_count}セルをマーク")
            return True

        except Exception as e:
            logger.error(f"Excelハイライトエラー: {e}")
            return False

    def _generate_highlighted_pdf_from_text(self, text: str, quotes: List[str], output_path: str, original_filename: str = "") -> bool:
        """テキストコンテンツからPDFを生成し、該当箇所をハイライト（背景色変更）"""
        try:
            # Ensure output_path is a Path object for name access
            output_path_obj = Path(output_path)

            # Use string path for reportlab canvas
            c = canvas.Canvas(str(output_path_obj), pagesize=A4)
            width, height = A4

            font_name = "Helvetica"
            font_size = 10
            line_height = 14
            margin = 50

            y = height - margin
            x = margin
            max_width = width - 2 * margin

            # ヘッダー
            c.setFont("Helvetica-Bold", 12)
            c.drawString(x, y, f"Expected Content from: {original_filename}")
            y -= 20

            c.setFont(font_name, font_size)

            lines = text.split('\n')
            normalized_quotes = [self._normalize_text(q) for q in quotes if q and len(q) >= 3]

            for line in lines:
                # 引用箇所が含まれているかチェック
                is_highlighted = False
                line_norm = self._normalize_text(line)

                # 通常のマッチ
                for quote in quotes:
                    if quote and len(quote) >= 3 and quote in line:
                        is_highlighted = True
                        break

                # 正規化マッチ
                if not is_highlighted:
                    for quote_norm in normalized_quotes:
                        if quote_norm in line_norm:
                            is_highlighted = True
                            break

                if is_highlighted:
                    c.setFillColor(yellow)
                    c.rect(x - 2, y - 2, max_width + 4, line_height, fill=1, stroke=0)
                    c.setFillColor(black)

                # 文字化け回避のため、ASCII文字のみ描画するか、置換して描画
                # ここでは簡易的にencode/decodeで非ASCIIを除外
                safe_line = line.encode('ascii', 'replace').decode('ascii')
                c.drawString(x, y + 2, safe_line[:100] + ("..." if len(safe_line) > 100 else ""))

                y -= line_height

                if y < margin:
                    c.showPage()
                    y = height - margin
                    c.setFont(font_name, font_size)

            c.save()
            logger.info(f"[TextPDF] PDF生成完了: {output_path_obj.name}")
            return True

        except Exception as e:
            logger.error(f"テキストPDF変換エラー: {e}")
            return False

def Stem(filename: str) -> str:
    """拡張子を除いたファイル名を取得"""
    return os.path.splitext(filename)[0]
