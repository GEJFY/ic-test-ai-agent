"""
================================================================================
highlighting_service.py - 証跡ハイライトサービス
================================================================================

【概要】
監査結果に基づいて証跡ファイルの該当箇所をハイライトします。
以下の3つのモードで動作します：
1. PDFハイライト: 既存のPDFファイル内のテキストをハイライト
2. Excelハイライト: Excelファイル内のセルをハイライト
3. テキスト変換PDF: Word/Email/CSV/画像などのテキストをPDF化してハイライト

【依存ライブラリ】
- fitz (PyMuPDF): PDF操作
- openpyxl: Excel操作
- reportlab: PDF生成

【使用例】
```python
service = HighlightingService()
highlighted_files = await service.highlight_evidence(audit_result, context)
```
"""
import os
import logging
import base64
import tempfile
import re
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

# 外部ライブラリのインポート（利用不可時のハンドリング付き）
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    openpyxl = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import yellow, black
except ImportError:
    canvas = None
    A4 = (595.27, 841.89)  # A4サイズのフォールバック値
    pdfmetrics = None
    TTFont = None
    yellow = None
    black = None

# from core.tasks.base_task import AuditContext, AuditResult, EvidenceFile
# 循環参照とパッケージ初期化の問題を回避するため core.types を使用
from core.types import AuditContext, AuditResult, EvidenceFile

logger = logging.getLogger(__name__)


class HighlightingService:
    """
    証跡ハイライト機能を提供するサービスクラス
    """

    def __init__(self, output_dir: str = "ic_audit_highlighted"):
        """
        初期化

        Args:
            output_dir (str): ハイライト済みファイルの保存先ディレクトリ名
                              クライアント側ではxlsmと同じフォルダ配下にこの名前で保存される
        """
        self.output_dir_name = output_dir
        self._check_dependencies()

    def _check_dependencies(self):
        """依存ライブラリの確認"""
        missing = []
        if not fitz:
            missing.append("pymupdf")
        if not openpyxl:
            missing.append("openpyxl")
        if not canvas:
            missing.append("reportlab")

        if missing:
            logger.warning(f"以下のライブラリが不足しています: {', '.join(missing)}。一部機能が制限されます。")
        else:
            logger.debug("ハイライト用ライブラリは全て利用可能です")

    def _normalize_text(self, text: str) -> str:
        """
        テキストの正規化（空白・改行の削除、全角・半角の統一）
        """
        if not text:
            return ""

        import unicodedata
        # NFKC正規化（全角英数→半角、半角カナ→全角など）
        text = unicodedata.normalize('NFKC', text)
        # 改行と空白を削除（厳密な一致ではなく、内容の一致を見るため）
        text = re.sub(r'[\s\u3000]+', '', text)
        return text

    async def highlight_evidence(self, audit_result: AuditResult, context: AuditContext) -> List[Dict[str, str]]:
        """
        監査結果に基づいて証跡ファイルをハイライト
        """
        if not audit_result.document_reference:
            logger.info("引用箇所がないためハイライトをスキップします")
            return []

        logger.info(f"ハイライト処理開始: {len(context.evidence_files)}ファイルの証跡")
        logger.info(f"[Highlight] document_reference (先頭200文字): {audit_result.document_reference[:200]}...")

        # 引用箇所を解析
        quotes_by_file = self._parse_quotes(audit_result.document_reference)
        logger.info(f"[Highlight] 抽出された引用: ファイル数={len(quotes_by_file)}, "
                    f"ファイル名={list(quotes_by_file.keys())}")
        for fname, qs in quotes_by_file.items():
            logger.info(f"[Highlight]   {fname}: {len(qs)}件の引用 "
                        f"(先頭: {qs[0][:50] if qs else '(なし)'}...)")

        # 全引用をフラット化（フォールバック用）
        all_quotes = []
        for qs in quotes_by_file.values():
            all_quotes.extend(qs)

        highlighted_files_info = []

        # 保存先ディレクトリの決定
        # サーバー側は一時ディレクトリで処理し、Base64でクライアントに返却する
        # クライアント（Excel VBA）がxlsmと同じフォルダのic_audit_highlightedに保存する
        save_dir = Path(tempfile.gettempdir()) / self.output_dir_name
        try:
            save_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.warning(f"保存先ディレクトリ作成失敗 ({save_dir}): {e}")
            save_dir = Path(tempfile.gettempdir()) / "ic_audit_highlighted"
            save_dir.mkdir(parents=True, exist_ok=True)

        for ef in context.evidence_files:
            try:
                # このファイルに関連する引用があるか確認
                file_quotes = []

                # ファイル名マッチングの強化 (正規化して比較)
                ef_stem = Stem(ef.file_name)
                ef_stem_norm = self._normalize_text(ef_stem)

                for fname, quotes in quotes_by_file.items():
                    # 1. 完全一致
                    if fname == ef.file_name:
                        file_quotes.extend(quotes)
                        logger.debug(f"[Highlight] ファイル名マッチ(完全一致): {fname} → {ef.file_name}")
                        continue

                    # 2. 部分一致 (相互包含)
                    if fname in ef.file_name or ef.file_name in fname:
                        file_quotes.extend(quotes)
                        logger.debug(f"[Highlight] ファイル名マッチ(部分一致): {fname} → {ef.file_name}")
                        continue

                    # 3. 拡張子なしでの一致
                    fname_stem = os.path.splitext(fname)[0]
                    if fname_stem == ef_stem:
                        file_quotes.extend(quotes)
                        logger.debug(f"[Highlight] ファイル名マッチ(stem一致): {fname} → {ef.file_name}")
                        continue

                    # 4. 正規化比較
                    fname_norm = self._normalize_text(fname)
                    if fname_norm in ef_stem_norm or ef_stem_norm in fname_norm:
                        file_quotes.extend(quotes)
                        logger.debug(f"[Highlight] ファイル名マッチ(正規化一致): {fname} → {ef.file_name}")

                # 重複排除
                file_quotes = list(set(file_quotes))

                # フォールバック: ファイル名マッチが0件かつ全引用がある場合、全引用を試行
                if not file_quotes and all_quotes:
                    logger.warning(f"[Highlight] ファイル名マッチ失敗: {ef.file_name} に一致する引用なし。"
                                   f"全引用({len(all_quotes)}件)をフォールバックとして適用します")
                    file_quotes = list(set(all_quotes))

                logger.info(f"ファイル処理: {ef.file_name} (抽出引用数: {len(file_quotes)})")

                output_filename = f"highlighted_{ef.file_name}"
                ext = ef.extension.lower()
                processed = False
                output_path = save_dir / output_filename

                # Base64デコードして一時ファイルに保存
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
                    tmp_file.write(base64.b64decode(ef.base64_content))
                    tmp_path = tmp_file.name

                if ext == '.pdf':
                    if fitz:
                        processed = self._highlight_pdf(tmp_path, file_quotes, str(output_path))
                    else:
                        processed = self._copy_file(tmp_path, output_path)

                elif ext in ['.xlsx', '.xls']:
                    if openpyxl:
                        if ext == '.xlsx':
                            processed = self._highlight_excel(tmp_path, file_quotes, str(output_path))
                        else:
                            logger.info(f"Excel (.xls) は直接ハイライト非対応のため、コピーのみを行います: {ef.file_name}")
                            processed = self._copy_file(tmp_path, output_path)
                    else:
                        processed = self._copy_file(tmp_path, output_path)

                else:
                    # その他の形式: テキストPDF化
                    content_text = ""
                    try:
                        with open(tmp_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content_text = f.read()
                    except Exception:
                        content_text = "(バイナリファイルのためテキスト表示不可)"

                    if canvas:
                        processed = self._generate_highlighted_pdf_from_text(
                            content_text, file_quotes, str(output_path), original_filename=ef.file_name
                        )
                    else:
                        # reportlabがない場合はコピー（拡張子そのまま）
                        processed = self._copy_file(tmp_path, output_path)

                # 一時ファイル削除
                try:
                    os.remove(tmp_path)
                except:
                    pass

                if processed:
                    # ファイルの内容をBase64で読み込む（クライアント側で保存するため）
                    try:
                        with open(output_path, "rb") as f:
                            file_content = f.read()
                            b64_content = base64.b64encode(file_content).decode("utf-8")
                    except Exception as e:
                        logger.error(f"ハイライト済みファイルの読み込み失敗: {e}")
                        b64_content = ""

                    highlighted_files_info.append({
                        "fileName": output_path.name,
                        "originalFileName": ef.file_name,  # 元のファイル名（表示用）
                        "filePath": str(output_path.parent),  # サーバー側パス（PowerShellがローカルパスに置換）
                        "base64": b64_content
                    })
                    logger.info(f"ハイライト済みファイル生成完了: {output_path.name}")

            except Exception as e:
                logger.error(f"ファイル処理エラー ({ef.file_name}): {e}", exc_info=True)

        return highlighted_files_info

    def _copy_file(self, src: str, dst: Path) -> bool:
        """ファイルをコピーする（ハイライトなしフォールバック）"""
        try:
            import shutil
            shutil.copy(src, dst)
            logger.debug(f"ファイルコピー完了（ハイライトなし）: {dst.name}")
            return True
        except Exception as e:
            logger.error(f"ファイルコピー失敗: {e}")
            return False

    def _parse_quotes(self, document_reference: str) -> Dict[str, List[str]]:
        """
        audit_result.document_reference からファイルごとの引用を抽出

        対応フォーマット:
        - [filename] location：quote_text (1行形式)
        - [filename] location：\nquote_text (複数行形式)
        """
        quotes = {}
        if not document_reference:
            return quotes

        text = document_reference.replace('\r\n', '\n')
        current_file = "unknown"

        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue

            # [filename] または 【filename】 の検出
            match = re.match(r'^[\[【](.*?)[\]】]', line)
            if match:
                current_file = match.group(1).strip()
                # ファイル検出時点でdict keyを必ず初期化
                if current_file not in quotes:
                    quotes[current_file] = []

                content = line[match.end():].strip()
                # プレフィックス除去
                content = re.sub(r'^(page|p\.?)\s*\d+.*?[：:]', '', content, flags=re.IGNORECASE).strip()
                content = re.sub(r'^.*?[：:]', '', content).strip()

                if content:
                    quotes[current_file].append(content)
            else:
                # 継続行: 前のファイルの引用テキスト
                if current_file != "unknown" and current_file in quotes:
                    quotes[current_file].append(line)

        return quotes

    def _split_quote_segments(self, quote: str, max_len: int = 30) -> List[str]:
        """
        長い引用テキストを検索用の短いセグメントに分割する。
        PDF内部の改行やフォーマットの違いに対応するため、
        短い単位で検索して各セグメントをハイライトする。
        """
        if len(quote) <= max_len:
            return [quote]

        segments = []
        # 句読点・句切りで分割を優先
        split_chars = ['。', '、', '．', '，', '. ', ', ', '；', ';', '\n']
        parts = [quote]
        for sep in split_chars:
            new_parts = []
            for part in parts:
                if len(part) > max_len:
                    sub = part.split(sep)
                    for i, s in enumerate(sub):
                        # 区切り文字を末尾に付与（最後以外）
                        if i < len(sub) - 1:
                            new_parts.append(s + sep)
                        elif s:
                            new_parts.append(s)
                else:
                    new_parts.append(part)
            parts = new_parts

        # それでも長いセグメントは文字数で強制分割
        for part in parts:
            part = part.strip()
            if not part or len(part) < 3:
                continue
            if len(part) <= max_len:
                segments.append(part)
            else:
                for i in range(0, len(part), max_len):
                    seg = part[i:i + max_len].strip()
                    if seg and len(seg) >= 3:
                        segments.append(seg)

        return segments if segments else [quote]

    def _highlight_pdf(self, input_path: str, quotes: List[str], output_path: str) -> bool:
        """PyMuPDFを使用してPDFをハイライト"""
        try:
            doc = fitz.open(input_path)
            found_count = 0

            # 有効な引用のみフィルタ
            valid_quotes = [q for q in quotes if q and len(q) >= 3]

            for page_num, page in enumerate(doc):
                page_text = page.get_text()
                page_text_norm = self._normalize_text(page_text)

                for quote in valid_quotes:
                    # 1. まず引用全体で検索
                    text_instances = page.search_for(quote)

                    if text_instances:
                        for inst in text_instances:
                            highlight = page.add_highlight_annot(inst)
                            highlight.set_colors(stroke=(1, 1, 0))
                            highlight.update()
                        found_count += len(text_instances)
                        continue

                    # 2. 全体検索で見つからない場合: セグメント分割で検索
                    quote_norm = self._normalize_text(quote)
                    if quote_norm in page_text_norm:
                        segments = self._split_quote_segments(quote)
                        for seg in segments:
                            seg_instances = page.search_for(seg)
                            if seg_instances:
                                for inst in seg_instances:
                                    highlight = page.add_highlight_annot(inst)
                                    highlight.set_colors(stroke=(1, 1, 0))
                                    highlight.update()
                                found_count += len(seg_instances)
                            else:
                                logger.debug(f"[PDF] ページ{page_num+1}: セグメント未検出: {seg[:20]}...")

            doc.save(output_path)
            doc.close()

            if found_count > 0:
                logger.info(f"[PDF] ハイライト完了: {found_count}箇所をマーク")
            else:
                logger.warning(f"[PDF] ハイライト0件: 検索した引用数={len(valid_quotes)}, "
                               f"引用サンプル={valid_quotes[0][:50] if valid_quotes else '(なし)'}...")
            return True

        except Exception as e:
            logger.error(f"PDFハイライトエラー: {e}")
            return False

    def _highlight_excel(self, input_path: str, quotes: List[str], output_path: str) -> bool:
        """openpyxlを使用してExcelセルをハイライト"""
        try:
            wb = openpyxl.load_workbook(input_path)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            found_count = 0

            # 正規化された引用リスト
            normalized_quotes = [self._normalize_text(q) for q in quotes if q and len(q) >= 3]

            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = str(cell.value)
                            cell_text_norm = self._normalize_text(cell_text)

                            matched = False
                            # 1. 通常の部分一致
                            for quote in quotes:
                                if quote in cell_text:
                                    matched = True
                                    break

                            # 2. 正規化部分一致
                            if not matched:
                                for quote_norm in normalized_quotes:
                                    if quote_norm in cell_text_norm:
                                        matched = True
                                        break

                            if matched:
                                cell.fill = yellow_fill
                                found_count += 1

            wb.save(output_path)
            if found_count > 0:
                logger.info(f"[Excel] ハイライト完了: {found_count}セルをマーク")
            else:
                logger.warning(f"[Excel] ハイライト0件: 検索した引用数={len(quotes)}, "
                               f"引用サンプル={quotes[0][:50] if quotes else '(なし)'}...")
            return True

        except Exception as e:
            logger.error(f"Excelハイライトエラー: {e}")
            return False

    def _register_japanese_font(self) -> Optional[str]:
        """
        日本語対応フォントを検出・登録する。
        Windows標準フォント → reportlab CIDフォントの順で試行。
        登録成功時はフォント名を返す。失敗時はNone。
        """
        # Windows標準日本語フォントのパス候補
        font_candidates = [
            ("MSGothic", r"C:\Windows\Fonts\msgothic.ttc"),
            ("MSGothic", r"C:\Windows\Fonts\msgothic.ttf"),
            ("YuGothic", r"C:\Windows\Fonts\YuGothR.ttc"),
            ("Meiryo", r"C:\Windows\Fonts\meiryo.ttc"),
            ("Meiryo", r"C:\Windows\Fonts\meiryo.ttf"),
        ]

        for font_name, font_path in font_candidates:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    logger.debug(f"[TextPDF] 日本語フォント登録成功: {font_name} ({font_path})")
                    return font_name
                except Exception as e:
                    logger.debug(f"[TextPDF] フォント登録失敗 ({font_path}): {e}")

        # reportlab CIDフォント(ビルトイン日本語フォント)をフォールバックとして試行
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
            logger.debug("[TextPDF] CIDフォント HeiseiMin-W3 を登録")
            return 'HeiseiMin-W3'
        except Exception as e:
            logger.debug(f"[TextPDF] CIDフォント登録失敗: {e}")

        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
            logger.debug("[TextPDF] CIDフォント HeiseiKakuGo-W5 を登録")
            return 'HeiseiKakuGo-W5'
        except Exception as e:
            logger.debug(f"[TextPDF] CIDフォント登録失敗: {e}")

        logger.warning("[TextPDF] 日本語フォントが見つかりません。ASCII文字のみ表示されます。")
        return None

    def _generate_highlighted_pdf_from_text(self, text: str, quotes: List[str], output_path: str, original_filename: str = "") -> bool:
        """テキストコンテンツからPDFを生成し、該当箇所をハイライト（背景色変更）"""
        if canvas is None:
            logger.warning("reportlabが未インストールのためPDF生成をスキップします")
            return False
        try:
            output_path_obj = Path(output_path)

            c = canvas.Canvas(str(output_path_obj), pagesize=A4)
            width, height = A4

            # 日本語フォントの登録を試行
            jp_font = self._register_japanese_font()
            font_name = jp_font if jp_font else "Helvetica"
            font_size = 10
            line_height = 14
            margin = 50

            y = height - margin
            x = margin
            max_width = width - 2 * margin

            # ヘッダー
            c.setFont(font_name, 12)
            header = f"Evidence: {original_filename}" if original_filename else "Evidence"
            c.drawString(x, y, header)
            y -= 20

            c.setFont(font_name, font_size)

            lines = text.split('\n')
            normalized_quotes = [self._normalize_text(q) for q in quotes if q and len(q) >= 3]

            for line in lines:
                is_highlighted = False
                line_norm = self._normalize_text(line)

                # 通常のマッチ
                for quote in quotes:
                    if quote and len(quote) >= 3 and quote in line:
                        is_highlighted = True
                        break

                # 正規化マッチ
                if not is_highlighted:
                    for quote_norm in normalized_quotes:
                        if quote_norm in line_norm:
                            is_highlighted = True
                            break

                if is_highlighted:
                    c.setFillColor(yellow)
                    c.rect(x - 2, y - 2, max_width + 4, line_height, fill=1, stroke=0)
                    c.setFillColor(black)

                # 日本語フォント利用時はそのまま描画、なければASCIIフォールバック
                if jp_font:
                    display_line = line[:100] + ("..." if len(line) > 100 else "")
                else:
                    display_line = line.encode('ascii', 'replace').decode('ascii')
                    display_line = display_line[:100] + ("..." if len(display_line) > 100 else "")
                c.drawString(x, y + 2, display_line)

                y -= line_height

                if y < margin:
                    c.showPage()
                    y = height - margin
                    c.setFont(font_name, font_size)

            c.save()
            logger.info(f"[TextPDF] PDF生成完了: {output_path_obj.name}")
            return True

        except Exception as e:
            logger.error(f"テキストPDF変換エラー: {e}")
            return False

def Stem(filename: str) -> str:
    """拡張子を除いたファイル名を取得"""
    return os.path.splitext(filename)[0]
